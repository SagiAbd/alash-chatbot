import tempfile
import unittest
from pathlib import Path

import openpyxl

from app.services.xlsx_processor import parse_glossary_xlsx


class XlsxProcessorTests(unittest.TestCase):
    def test_parse_glossary_xlsx_finds_terms_on_non_active_sheet(self) -> None:
        """Parser should scan all sheets and ignore shifting metadata rows."""

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "terms.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Notes"
            sheet = workbook.create_sheet("Terms")
            sheet.append(["Кітап атауы", "Book"])
            sheet.append(["Авторы", "Author"])
            sheet.append(["Алаш термині", "Сала", "Авторы"])
            sheet.append(["Ұлт", "Тарих", "Ахмет"])
            workbook.save(path)
            workbook.close()

            terms = parse_glossary_xlsx(str(path))

        self.assertEqual(1, len(terms))
        self.assertEqual("Ұлт", terms[0]["alash_term"])
        self.assertEqual("Тарих", terms[0]["field"])
        self.assertEqual("Ахмет", terms[0]["author"])

    def test_parse_glossary_xlsx_allows_removed_optional_columns(self) -> None:
        """Parser should accept reduced exports as long as the term column exists."""

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "terms.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Terms"
            sheet.append(["Алаш термині", "Сілтеме"])
            sheet.append(["Алаш", "https://example.test/term"])
            workbook.save(path)
            workbook.close()

            terms = parse_glossary_xlsx(str(path))

        self.assertEqual(1, len(terms))
        self.assertEqual("Алаш", terms[0]["alash_term"])
        self.assertEqual("https://example.test/term", terms[0]["link"])
        self.assertIn("Сілтеме: https://example.test/term", terms[0]["page_content"])

    def test_parse_glossary_xlsx_rejects_invalid_workbooks(self) -> None:
        """Invalid files should raise a clean validation error."""

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "broken.xlsx"
            path.write_text("not really an xlsx", encoding="utf-8")

            with self.assertRaisesRegex(
                ValueError, "not a valid Excel workbook"
            ):
                parse_glossary_xlsx(str(path))


if __name__ == "__main__":
    unittest.main()
