"""Parse Alash-era scientific term glossary xlsx files into term chunks.

Detects the header row by column name rather than position, so the parser
is resilient to columns being added, removed, or reordered between exports.
"""

import logging
import re
import unicodedata
from typing import Any
from zipfile import BadZipFile

logger = logging.getLogger(__name__)

# (substring_to_match_in_header_cell, normalized_field_key)
# Lower-cased substring matching — resilient to minor wording differences.
_COLUMN_PATTERNS: list[tuple[str, str]] = [
    ("алаш термині", "alash_term"),
    ("заманауи термин", "modern_term"),
    ("заманауи атауы", "modern_term"),
    ("сала", "field"),
    ("кіші сала", "field"),
    ("заманауи түсініктеме", "modern_definition"),
    ("алаш түсініктемесі", "alash_definition"),
    ("анықтама бар ма", "has_definition"),
    ("екі бет арасындағы мәтін", "context"),
    ("контекст", "context"),
    ("авторы", "author"),
    ("автор", "author"),
    ("басталатын беті", "start_page"),
    ("аяқталу беті", "end_page"),
    ("жазылу жылы", "year"),
    ("сілтеме", "link"),
]

_METADATA_LABELS: dict[str, str] = {
    "кітап атауы": "title",
    "авторы": "author",
    "жазылу жылы": "year",
    "сілтеме": "link",
}


def _normalize_cell_text(value: str) -> str:
    """Normalize header text for resilient substring matching."""
    normalized = unicodedata.normalize("NFKC", value or "")
    normalized = normalized.replace("\ufeff", " ").replace("\xa0", " ")
    normalized = normalized.strip().lower()
    return re.sub(r"\s+", " ", normalized)


def _detect_column_mapping(row: list[str]) -> dict[int, str]:
    """Return {column_index: field_key} by scanning a header row."""
    mapping: dict[int, str] = {}
    for col_idx, cell in enumerate(row):
        normalized = _normalize_cell_text(cell)
        if not normalized:
            continue
        for pattern, key in _COLUMN_PATTERNS:
            if pattern in normalized:
                if key not in mapping.values():  # first match wins
                    mapping[col_idx] = key
                break
    return mapping


def _build_page_content(term: dict[str, Any]) -> str:
    """Build a single searchable text blob from all non-empty term fields."""
    parts = []
    for key, label in [
        ("alash_term", "Алаш термині"),
        ("modern_term", "Заманауи термин"),
        ("field", "Сала"),
        ("author", "Автор"),
        ("modern_definition", "Анықтама"),
        ("alash_definition", "Алаш анықтамасы"),
        ("context", "Контекст"),
    ]:
        val = str(term.get(key) or "").strip()
        if val:
            parts.append(f"{label}: {val}")
    for key, label in [
        ("year", "Жылы"),
        ("link", "Сілтеме"),
        ("start_page", "Басталатын беті"),
        ("end_page", "Аяқталу беті"),
    ]:
        val = str(term.get(key) or "").strip()
        if val:
            parts.append(f"{label}: {val}")
    return " | ".join(parts)


def _extract_sheet_metadata(ws: Any) -> dict[str, str]:
    """Extract the workbook-style metadata header block from a worksheet."""
    metadata: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        first = ""
        second = ""
        if row and len(row) > 0 and row[0] is not None:
            first = str(row[0]).strip()
        if row and len(row) > 1 and row[1] is not None:
            second = str(row[1]).strip()
        normalized_label = _normalize_cell_text(first)
        field_key = _METADATA_LABELS.get(normalized_label)
        if field_key and second:
            metadata[field_key] = second
    return metadata


def _extract_terms_from_sheet(ws: Any) -> list[dict[str, Any]]:
    """Extract terms from a single worksheet."""
    col_mapping: dict[int, str] = {}
    terms: list[dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        row_values = [str(cell) if cell is not None else "" for cell in row]

        if not col_mapping:
            candidate = _detect_column_mapping(row_values)
            if "alash_term" in candidate.values():
                col_mapping = candidate
            continue

        term: dict[str, Any] = {}
        for col_idx, field_key in col_mapping.items():
            if col_idx < len(row_values):
                val = row_values[col_idx].strip()
                term[field_key] = val if val and val.lower() != "none" else ""

        alash_term = str(term.get("alash_term") or "").strip()
        if not alash_term:
            continue

        term["page_content"] = _build_page_content(term)
        terms.append(term)

    return terms


def parse_glossary_xlsx(
    file_path: str,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Parse a glossary xlsx and return term dicts plus parsed sheet metadata.

    Each dict contains normalized field keys and a ``page_content`` key
    suitable for full-text search.  The header row is detected by column
    names, not by position, so the file layout can change without breaking
    parsing as long as at least three recognisable columns remain.

    Args:
        file_path: Local path to the xlsx file.

    Returns:
        Tuple of parsed term dicts and metadata from the workbook header block.
        Metadata may include ``title``, ``author``, ``year``, ``link``, and
        ``sheet_title``. Returns ``([], {})`` if no valid header was found.
    """
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValueError(
            "The uploaded file is not a valid Excel workbook."
        ) from exc

    try:
        for ws in wb.worksheets:
            terms = _extract_terms_from_sheet(ws)
            if terms:
                metadata = _extract_sheet_metadata(ws)
                metadata["sheet_title"] = ws.title
                logger.info(
                    "Parsed %d terms from %s (sheet=%s)",
                    len(terms),
                    file_path,
                    ws.title,
                )
                return terms, metadata
    finally:
        wb.close()

    logger.info("Parsed 0 terms from %s", file_path)
    return [], {}
